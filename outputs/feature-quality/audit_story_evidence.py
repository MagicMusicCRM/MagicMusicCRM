from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
QUALITY = ROOT / "outputs" / "feature-quality"
WORKBOOK = QUALITY / "MagicMusicCRM-Canonical-Feature-Register.xlsx"

def relative_file(value: str | None) -> str:
    if not value:
        return ""
    if value.startswith("file:"):
        value = unquote(urlparse(value).path).lstrip("/")
    value = value.replace("\\", "/")
    marker = "/MagicMusicCRM/"
    return value.split(marker, 1)[-1] if marker in value else value


def flutter_tests() -> list[dict]:
    source = QUALITY / "evidence" / "flutter-test-machine-final-156.ndjson"
    suites: dict[int, str] = {}
    starts: dict[int, dict] = {}
    done: dict[int, dict] = {}
    with source.open(encoding="utf-8-sig", errors="replace") as handle:
        for line in handle:
            line = line.strip()
            if not line.startswith("{"):
                continue
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                continue
            if event.get("type") == "suite":
                suite = event["suite"]
                suites[suite["id"]] = relative_file(suite.get("path"))
            elif event.get("type") == "testStart":
                test = event["test"]
                starts[test["id"]] = test
            elif event.get("type") == "testDone":
                done[event["testID"]] = {
                    "result": event.get("result", "unknown"),
                    "hidden": event.get("hidden", False),
                }
    return [
        {
            "platform": "Flutter",
            "file": relative_file(test.get("root_url")) or suites.get(test["suiteID"], ""),
            "name": test["name"],
            "result": done.get(test_id, {}).get("result", "unknown"),
            "line": test.get("root_line") or test.get("line"),
        }
        for test_id, test in starts.items()
        if not done.get(test_id, {}).get("hidden", False)
    ]


def server_tests() -> list[dict]:
    source = QUALITY / "evidence" / "server-test-json-final-156.json"
    payload = json.loads(source.read_text(encoding="utf-8"))
    result: list[dict] = []
    for suite in payload["testResults"]:
        file = relative_file(suite["name"])
        for assertion in suite["assertionResults"]:
            result.append(
                {
                    "platform": "Server",
                    "file": file,
                    "name": assertion["fullName"],
                    "result": "success" if assertion["status"] == "passed" else assertion["status"],
                    "line": None,
                }
            )
    return result


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=False)
    stories_sheet = workbook["User Stories"]
    stories = [
        {
            "id": stories_sheet.cell(row, 1).value,
            "system": stories_sheet.cell(row, 2).value,
            "feature": stories_sheet.cell(row, 3).value,
            "expected": stories_sheet.cell(row, 8).value,
        }
        for row in range(2, stories_sheet.max_row + 1)
    ]

    runs_sheet = workbook["Test Runs"]
    device_runs: dict[str, list[dict]] = {}
    for row in range(2, runs_sheet.max_row + 1):
        platform = str(runs_sheet.cell(row, 6).value or "")
        if "Android" not in platform and "Windows" not in platform:
            continue
        story_id = runs_sheet.cell(row, 2).value
        device_runs.setdefault(story_id, []).append(
            {
                "run_id": runs_sheet.cell(row, 1).value,
                "phase": runs_sheet.cell(row, 3).value,
                "platform": platform,
                "result": runs_sheet.cell(row, 9).value,
                "evidence": runs_sheet.cell(row, 11).value,
                "observed": runs_sheet.cell(row, 10).value,
            }
        )

    tests = flutter_tests() + server_tests()
    rules = json.loads(
        (QUALITY / "story-evidence-rules.json").read_text(encoding="utf-8")
    )
    verified = []
    for story in stories:
        runs = device_runs.get(story["id"], [])
        latest_device = runs[-1] if runs else None
        selector_results = []
        for selector in rules.get(story["id"], []):
            file_part, name_part = selector.split("::", 1)
            matches = [
                test
                for test in tests
                if test["result"] == "success"
                and file_part.lower() in test["file"].lower()
                and name_part.lower() in test["name"].lower()
            ]
            selector_results.append(
                {"selector": selector, "matched": bool(matches), "tests": matches}
            )
        device_pass = latest_device is not None and latest_device["result"] == "PASS"
        automated_pass = bool(selector_results) and all(
            result["matched"] for result in selector_results
        )
        verified.append(
            {
                **story,
                "verified": automated_pass,
                "evidence_type": "AUTOMATED + DEVICE" if device_pass else "AUTOMATED",
                "baseline_device_pass": device_pass,
                "latest_device_run": latest_device,
                "selectors": selector_results,
            }
        )

    (QUALITY / "story-evidence-audit.json").write_text(
        json.dumps(verified, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    gaps = [item["id"] for item in verified if not item["verified"]]
    print(
        json.dumps(
            {
                "stories": len(stories),
                "flutter_tests": sum(test["platform"] == "Flutter" for test in tests),
                "server_tests": sum(test["platform"] == "Server" for test in tests),
                "stories_with_device_runs": sum(bool(runs) for runs in device_runs.values()),
                "stories_verified": len(verified) - len(gaps),
                "story_evidence_gaps": gaps,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    if gaps:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
